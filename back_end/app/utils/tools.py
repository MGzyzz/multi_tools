from urllib.parse import quote

import openpyxl
from django.core.cache import cache
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from app.models import Attendance, AttendanceStat, Group
from app.models._choices.attendanceChoices import AttendanceStatusChoices

# Minimum seconds between two export requests from the same teacher. Prevents a
# burst of heavy spreadsheet generations (e.g. accidental double-clicks or many
# tabs) from tying up worker threads. Kept synchronous on purpose — a single
# teacher's dataset is small and generates in well under a second.
EXPORT_RATE_LIMIT_SECONDS = 10

HEADER_FONT = Font(bold=True)

# Symbols used in the journal sheet for each attendance status.
STATUS_SYMBOL = {
    AttendanceStatusChoices.PRESENT.value: "+",
    AttendanceStatusChoices.LATE.value: "О",
    AttendanceStatusChoices.ABSENT.value: "Н",
    AttendanceStatusChoices.NOT_MARKED.value: "",
}
JOURNAL_LEGEND = (
    "Условные обозначения:  + присутствовал   О опоздал   " "Н отсутствовал   (пусто) не отмечено"
)


def _attendance_percent(attended: int, total: int) -> int:
    """Attendance percentage, capped at 100% (AttendanceStat.attended can drift
    above total due to signal desync)."""
    if not total:
        return 0
    return min(100, round(attended / total * 100))


def _style_header(worksheet, columns):
    """Write a bold header row and set sensible column widths."""
    worksheet.append([title for title, _ in columns])
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
    for index, (_, width) in enumerate(columns, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _fill_journal(worksheet, groups, group_ids):
    """
    Per-group attendance journal: students as rows, lessons (by date) as columns,
    with a status symbol in each cell — so it's clear on which day a student was
    absent. Multiple groups are stacked as separate blocks.
    """
    rows = (
        Attendance.objects.filter(schedule__group_id__in=group_ids)
        .values("schedule__group_id", "schedule_id", "schedule__date", "student_id", "status")
        .order_by("schedule__date", "schedule__time")
    )

    group_lessons: dict[int, list[tuple]] = {}  # group_id -> ordered [(schedule_id, date)]
    seen_lessons: dict[int, set] = {}
    status_map: dict[tuple, str] = {}  # (group_id, student_id, schedule_id) -> status
    for row in rows:
        gid = row["schedule__group_id"]
        sched_id = row["schedule_id"]
        group_lessons.setdefault(gid, [])
        seen_lessons.setdefault(gid, set())
        if sched_id not in seen_lessons[gid]:
            seen_lessons[gid].add(sched_id)
            group_lessons[gid].append((sched_id, row["schedule__date"]))
        status_map[(gid, row["student_id"], sched_id)] = row["status"]

    worksheet.append([JOURNAL_LEGEND])
    worksheet["A1"].font = Font(italic=True)

    max_cols = 1
    for group in groups:
        lessons = group_lessons.get(group.id, [])
        max_cols = max(max_cols, len(lessons) + 1)

        worksheet.append([])  # blank separator row
        worksheet.append([group.name])
        worksheet.cell(row=worksheet.max_row, column=1).font = HEADER_FONT

        worksheet.append(["Студент"] + [date.strftime("%d.%m") for _sid, date in lessons])
        for cell in worksheet[worksheet.max_row]:
            cell.font = HEADER_FONT

        for student in group.students.all():
            line = [f"{student.first_name} {student.last_name}".strip()]
            for sched_id, _date in lessons:
                line.append(STATUS_SYMBOL.get(status_map.get((group.id, student.id, sched_id)), ""))
            worksheet.append(line)

    worksheet.column_dimensions["A"].width = 28
    for col in range(2, max_cols + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 7


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def create_excel_attendance_file(request):
    """
    Export the authenticated teacher's groups, students and attendance stats as a
    multi-sheet Excel file.

    Scoped to ``request.user.teacher_profile`` — a teacher can only export their
    own data. Pass ``?group_id=N`` to export a single group (must belong to the
    teacher); otherwise all of the teacher's groups are exported.
    """
    teacher = getattr(request.user, "teacher_profile", None)
    if teacher is None:
        return Response(
            {"detail": "Only teachers can export attendance data."},
            status=status.HTTP_403_FORBIDDEN,
        )

    # Per-teacher rate limit (atomic: cache.add only succeeds if the key is absent).
    rate_key = f"export_rate:{teacher.id}"
    if not cache.add(rate_key, 1, timeout=EXPORT_RATE_LIMIT_SECONDS):
        return Response(
            {"detail": "Export was requested too recently. Please wait a few seconds."},
            status=status.HTTP_429_TOO_MANY_REQUESTS,
        )

    groups = Group.objects.filter(teacher=teacher).prefetch_related("students")

    group_id = request.query_params.get("group_id")
    if group_id:
        try:
            group_id = int(group_id)
        except (TypeError, ValueError):
            return Response(
                {"detail": "group_id must be an integer."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        groups = groups.filter(id=group_id)
        if not groups.exists():
            return Response(
                {"detail": "Group not found or not owned by this teacher."},
                status=status.HTTP_404_NOT_FOUND,
            )

    groups = list(groups.order_by("course", "name"))
    group_ids = [group.id for group in groups]

    workbook = openpyxl.Workbook()

    # --- Sheet 1: Groups -----------------------------------------------------
    groups_sheet = workbook.active
    groups_sheet.title = "Группы"
    _style_header(
        groups_sheet,
        [("Группа", 24), ("Курс", 10), ("Специальность", 30), ("Студентов", 12)],
    )
    for group in groups:
        groups_sheet.append(
            [
                group.name,
                group.course,
                group.group_specialty,
                group.students.count(),
            ]
        )

    # --- Sheet 2: Students ---------------------------------------------------
    students_sheet = workbook.create_sheet("Студенты")
    _style_header(
        students_sheet,
        [("Студент", 28), ("Группа", 24), ("Email", 28), ("Telegram", 20)],
    )
    for group in groups:
        for student in group.students.all():
            students_sheet.append(
                [
                    f"{student.first_name} {student.last_name}".strip(),
                    group.name,
                    student.email or "—",
                    student.telegram_username or "—",
                ]
            )

    # --- Sheet 3: Attendance stats ------------------------------------------
    attendance_sheet = workbook.create_sheet("Посещаемость")
    _style_header(
        attendance_sheet,
        [
            ("Студент", 28),
            ("Группа", 24),
            ("Предмет", 28),
            ("Посещено", 12),
            ("Всего", 10),
            ("Процент", 12),
        ],
    )
    stats = (
        AttendanceStat.objects.filter(group_id__in=group_ids)
        .select_related("student", "subject", "group")
        .order_by("group__name", "student__last_name", "subject__name")
    )
    for stat in stats:
        attendance_sheet.append(
            [
                f"{stat.student.first_name} {stat.student.last_name}".strip(),
                stat.group.name,
                stat.subject.name,
                min(stat.attended, stat.total),
                stat.total,
                f"{_attendance_percent(stat.attended, stat.total)}%",
            ]
        )

    # --- Sheet 4: Attendance journal (per-day) ------------------------------
    journal_sheet = workbook.create_sheet("Журнал")
    _fill_journal(journal_sheet, groups, group_ids)

    # --- Build response ------------------------------------------------------
    stamp = timezone.localdate().isoformat()
    if len(groups) == 1:
        safe_name = groups[0].name.replace(" ", "_")
        filename = f"lectern_{safe_name}_{stamp}.xlsx"
    else:
        filename = f"lectern_attendance_{stamp}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # RFC 5987: keep the header pure-ASCII (filename* carries the UTF-8 name,
    # percent-encoded) so Django does not fall back to RFC 2047 word-encoding,
    # which browsers don't understand for downloads.
    response["Content-Disposition"] = (
        f'attachment; filename="lectern_attendance.xlsx"; ' f"filename*=UTF-8''{quote(filename)}"
    )
    workbook.save(response)
    return response
